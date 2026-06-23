from __future__ import annotations

import asyncio
import calendar
import logging
import sys
from dataclasses import dataclass
from datetime import date, datetime

try:
    import asyncpg
except ModuleNotFoundError:  # pragma: no cover - lets pure unit tests import helpers before deps are installed.
    asyncpg = None  # type: ignore[assignment]

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:  # pragma: no cover - checked when Excel sync runs.
    load_workbook = None  # type: ignore[assignment]

from slack_sdk.web.async_client import AsyncWebClient

import db
from config import Settings, load_settings
from utils import slack_error_reason

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class BirthdayRow:
    email: str
    birth_month: int
    birth_day: int


@dataclass(frozen=True)
class SyncResult:
    upserted: int
    deactivated: int
    skipped: int = 0


@dataclass(frozen=True)
class ExcelSyncResult:
    upserted: int
    skipped: int


@dataclass(frozen=True)
class SlackLookupResult:
    slack_user_id: str | None = None
    abort_batch: bool = False


async def sync_hr_sheet(
    *, pool: asyncpg.Pool, client: AsyncWebClient, settings: Settings
) -> SyncResult:
    rows, skipped = read_excel_rows(settings.hr_excel_path)
    active_slack_user_ids: list[str] = []
    upserted = 0

    for row in rows:
        lookup = await resolve_slack_id_for_batch(client, row.email)
        if lookup.abort_batch:
            logger.error(
                "Aborting HR sheet sync: %s upserted, %s skipped before batch abort",
                upserted,
                skipped,
            )
            return SyncResult(upserted=upserted, deactivated=0, skipped=skipped)

        slack_user_id = lookup.slack_user_id
        if slack_user_id is None:
            logger.warning("Skipping HR row with unmatched email: %s", redact_email(row.email))
            skipped += 1
            continue

        await db.upsert_birthday(
            pool,
            slack_user_id=slack_user_id,
            birth_month=row.birth_month,
            birth_day=row.birth_day,
            email=row.email,
            source="hr",
        )
        active_slack_user_ids.append(slack_user_id)
        upserted += 1

    if active_slack_user_ids:
        deactivated = await db.mark_missing_birthdays_inactive(pool, active_slack_user_ids)
    else:
        logger.warning("No Slack users resolved from HR sheet; skipping soft-delete step")
        deactivated = 0

    return SyncResult(upserted=upserted, deactivated=deactivated, skipped=skipped)


def sync_from_excel(file_path: str) -> ExcelSyncResult:
    return asyncio.run(_sync_from_excel(file_path))


async def _sync_from_excel(file_path: str) -> ExcelSyncResult:
    rows, skipped = read_excel_rows(file_path)
    settings = load_settings()
    pool = await db.create_pool(settings.database_url)
    client = AsyncWebClient(token=settings.slack_bot_token)

    try:
        upserted = 0
        for row in rows:
            lookup = await resolve_slack_id_for_batch(client, row.email)
            if lookup.abort_batch:
                logger.error(
                    "Aborting Excel sync: %s upserted, %s skipped before batch abort",
                    upserted,
                    skipped,
                )
                return ExcelSyncResult(upserted=upserted, skipped=skipped)

            slack_user_id = lookup.slack_user_id
            if slack_user_id is None:
                logger.warning("Skipping Excel row with unmatched email: %s", redact_email(row.email))
                skipped += 1
                continue

            await db.upsert_birthday(
                pool,
                slack_user_id=slack_user_id,
                birth_month=row.birth_month,
                birth_day=row.birth_day,
                email=row.email,
                source="hr",
            )
            upserted += 1
    finally:
        await pool.close()

    return ExcelSyncResult(upserted=upserted, skipped=skipped)


async def resolve_slack_id(client: AsyncWebClient, email: str) -> str | None:
    return (await resolve_slack_id_for_batch(client, email)).slack_user_id


async def resolve_slack_id_for_batch(client: AsyncWebClient, email: str) -> SlackLookupResult:
    try:
        return SlackLookupResult(slack_user_id=await lookup_slack_user_by_email(client, email))
    except Exception as error:
        return await handle_slack_lookup_error(client, email, error)


async def lookup_slack_user_by_email(client: AsyncWebClient, email: str) -> str | None:
    result = await client.users_lookupByEmail(email=email)
    user = result.get("user") or {}
    return user.get("id")


async def handle_slack_lookup_error(
    client: AsyncWebClient, email: str, error: Exception
) -> SlackLookupResult:
    reason = slack_error_reason(error)
    if reason == "users_not_found":
        logger.warning("Skipping Slack lookup for %s: users_not_found", redact_email(email))
        return SlackLookupResult()

    if reason == "ratelimited":
        retry_after = retry_after_seconds(error)
        logger.error(
            "Slack lookup for %s was rate limited; retrying after %s seconds",
            redact_email(email),
            retry_after,
        )
        await asyncio.sleep(retry_after)
        return await retry_slack_lookup_once(client, email)

    logger.error("Slack lookup for %s failed: %s", redact_email(email), reason, exc_info=True)
    return SlackLookupResult(abort_batch=True)


async def retry_slack_lookup_once(client: AsyncWebClient, email: str) -> SlackLookupResult:
    try:
        return SlackLookupResult(slack_user_id=await lookup_slack_user_by_email(client, email))
    except Exception as error:
        reason = slack_error_reason(error)
        if reason == "users_not_found":
            logger.warning(
                "Skipping Slack lookup for %s after retry: users_not_found",
                redact_email(email),
            )
            return SlackLookupResult()

        logger.error(
            "Slack lookup retry for %s failed: %s",
            redact_email(email),
            reason,
            exc_info=True,
        )
        return SlackLookupResult(abort_batch=True)


def retry_after_seconds(error: SlackApiError) -> int:
    response = getattr(error, "response", None)
    headers = getattr(response, "headers", None)
    if headers is None and isinstance(response, dict):
        headers = response.get("headers")

    if headers is not None:
        retry_after = header_value(headers, "Retry-After")
        if retry_after is not None:
            try:
                return max(0, int(retry_after))
            except ValueError:
                pass

    return 1


def header_value(headers: object, name: str) -> str | None:
    if hasattr(headers, "get"):
        value = headers.get(name) or headers.get(name.lower())
        return str(value) if value is not None else None

    return None


def read_excel_rows(file_path: str) -> tuple[list[BirthdayRow], int]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to read Excel files")

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        rows: list[BirthdayRow] = []
        skipped = 0
        birthday_index = 0
        email_index = 2
        columns_detected = False
        for index, raw_row in enumerate(workbook.active.iter_rows(values_only=True), start=1):
            if not columns_detected:
                header_indexes = _excel_header_indexes(raw_row)
                if header_indexes is not None:
                    birthday_index, email_index = header_indexes
                    columns_detected = True
                    continue
                columns_detected = True

            if not any(raw_row):
                continue

            email = _excel_cell_text(_excel_row_value(raw_row, email_index))
            birthday_value = _excel_row_value(raw_row, birthday_index)
            birthday = _excel_birthday_mm_dd(birthday_value)
            if not email:
                logger.warning("Skipping Excel row %s with missing email", index)
                skipped += 1
                continue

            if birthday is None:
                logger.warning(
                    "Skipping Excel row %s with invalid birthday: %s",
                    index,
                    _excel_cell_text(birthday_value),
                )
                skipped += 1
                continue

            birth_month, birth_day = _parse_mm_dd(birthday)
            rows.append(BirthdayRow(email=email, birth_month=birth_month, birth_day=birth_day))

        return rows, skipped
    finally:
        workbook.close()


def parse_rows(values: list[list[str]]) -> list[BirthdayRow]:
    if not values:
        return []

    headers = [cell.strip().lower() for cell in values[0]]
    rows: list[BirthdayRow] = []
    for raw_row in values[1:]:
        row = _row_dict(headers, raw_row)
        email = (row.get("email") or "").strip()
        if not email:
            continue

        birthday = _parse_birthday(row)
        if birthday is None:
            logger.warning("Skipping HR row with invalid birthday: %s", email)
            continue

        birth_month, birth_day = birthday
        rows.append(BirthdayRow(email=email, birth_month=birth_month, birth_day=birth_day))

    return rows


def _row_dict(headers: list[str], row: list[str]) -> dict[str, str]:
    return {header: row[index] if index < len(row) else "" for index, header in enumerate(headers)}


def _excel_cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _excel_row_value(row: tuple[object, ...], index: int) -> object | None:
    return row[index] if index < len(row) else None


def _excel_header_indexes(row: tuple[object, ...]) -> tuple[int, int] | None:
    headers = [_excel_cell_text(value).lower() for value in row]
    birthday_names = {"birthday", "birthdate", "생일", "생년월일"}
    email_names = {"email", "이메일"}

    birthday_index = next(
        (index for index, header in enumerate(headers) if header in birthday_names), None
    )
    email_index = next((index for index, header in enumerate(headers) if header in email_names), None)
    if birthday_index is None or email_index is None:
        return None
    return birthday_index, email_index


def _excel_birthday_mm_dd(value: object) -> str | None:
    if isinstance(value, (datetime, date)):
        return f"{value.month:02d}-{value.day:02d}"

    birthday = _parse_birthday_text(_excel_cell_text(value))
    if birthday is None:
        return None

    month, day = birthday
    return f"{month:02d}-{day:02d}"


def _parse_mm_dd(value: str) -> tuple[int, int] | None:
    parts = value.split("-", maxsplit=1)
    if len(parts) != 2:
        return None
    return _valid_month_day(parts[0], parts[1])


def _parse_birthday(row: dict[str, str]) -> tuple[int, int] | None:
    month = row.get("birth_month") or row.get("month")
    day = row.get("birth_day") or row.get("day")
    if month and day:
        return _valid_month_day(month, day)

    birthday = row.get("birthday") or row.get("birthdate")
    if not birthday:
        return None

    return _parse_birthday_text(birthday)


def _parse_birthday_text(value: str) -> tuple[int, int] | None:
    birthday = value.strip()
    if not birthday:
        return None

    parts = birthday.split("-")
    if len(parts) == 3:
        if len(parts[0]) == 4:
            return _valid_month_day(parts[1], parts[2])
        return _valid_month_day(parts[0], parts[1])

    if len(parts) == 2:
        return _valid_month_day(parts[0], parts[1])

    parts = birthday.split("/")
    if len(parts) == 2:
        return _valid_month_day(parts[0], parts[1])

    return None


def _valid_month_day(month: str, day: str) -> tuple[int, int] | None:
    try:
        month_int = int(month)
        day_int = int(day)
    except ValueError:
        return None

    if 1 <= month_int <= 12 and 1 <= day_int <= calendar.monthrange(2024, month_int)[1]:
        return month_int, day_int
    return None


def redact_email(email: str) -> str:
    local, separator, domain = email.partition("@")
    if not separator:
        return "[redacted]"
    visible = local[:1] if local else ""
    return f"{visible}***@{domain}"


def main(argv: list[str] | None = None) -> int:
    logging.basicConfig(level=logging.INFO)
    args = sys.argv[1:] if argv is None else argv
    if not args:
        raise SystemExit("Usage: python sync.py <excel-file>")

    result = sync_from_excel(args[0])
    print(f"Excel sync complete: {result.upserted} upserted, {result.skipped} skipped")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
