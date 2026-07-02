import asyncio
import sys
import types
from datetime import datetime
from types import SimpleNamespace

from openpyxl import Workbook


errors = types.ModuleType("slack_sdk.errors")


class SlackApiError(Exception):
    pass


errors.SlackApiError = SlackApiError
async_client = types.ModuleType("slack_sdk.web.async_client")
async_client.AsyncWebClient = object
sys.modules["slack_sdk"] = types.ModuleType("slack_sdk")
sys.modules["slack_sdk.errors"] = errors
sys.modules["slack_sdk.web"] = types.ModuleType("slack_sdk.web")
sys.modules["slack_sdk.web.async_client"] = async_client

from sync import _excel_birthday_mm_dd, parse_rows, read_excel_rows, sync_hr_sheet


def run(coro):
    return asyncio.run(coro)


def test_excel_birthday_mm_dd_yy_uses_month_day_prefix() -> None:
    assert _excel_birthday_mm_dd("03-21-91") == "03-21"


def test_excel_birthday_formats_use_supported_month_day_values(tmp_path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["birthday", "email"])
    sheet.append(["1991-03-21", "year@example.com"])
    sheet.append(["03/22", "slash@example.com"])
    sheet.append(["invalid", "bad@example.com"])
    file_path = tmp_path / "employees.xlsx"
    workbook.save(file_path)

    rows, skipped = read_excel_rows(str(file_path))

    assert skipped == 1
    assert [(row.email, row.birth_month, row.birth_day) for row in rows] == [
        ("year@example.com", 3, 21),
        ("slash@example.com", 3, 22),
    ]


def test_read_excel_rows_skips_formula_like_cells(tmp_path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["birthday", "email"])
    sheet.append(["03-21", "=HYPERLINK(\"https://example.com\")"])
    sheet.append(["=DATE(2026,3,22)", "formula-birthday@example.com"])
    sheet.append(["03-23", "safe@example.com"])
    file_path = tmp_path / "employees.xlsx"
    workbook.save(file_path)

    rows, skipped = read_excel_rows(str(file_path))

    assert skipped == 2
    assert [(row.email, row.birth_month, row.birth_day) for row in rows] == [
        ("safe@example.com", 3, 23),
    ]


def test_parse_rows_accepts_documented_birthday_formats() -> None:
    rows = parse_rows(
        [
            ["birthday", "email"],
            ["03-21-91", "dash-year@example.com"],
            ["1991-03-22", "year@example.com"],
            ["03/23", "slash@example.com"],
            ["invalid", "bad@example.com"],
        ]
    )

    assert [(row.email, row.birth_month, row.birth_day) for row in rows] == [
        ("dash-year@example.com", 3, 21),
        ("year@example.com", 3, 22),
        ("slash@example.com", 3, 23),
    ]


def test_parse_rows_skips_formula_like_cells() -> None:
    rows = parse_rows(
        [
            ["birthday", "email"],
            ["03-21", "@attacker@example.com"],
            ["=DATE(2026,3,22)", "formula-birthday@example.com"],
            ["03-23", "safe@example.com"],
        ]
    )

    assert [(row.email, row.birth_month, row.birth_day) for row in rows] == [
        ("safe@example.com", 3, 23),
    ]


def test_read_excel_rows_detects_employee_roster_headers(tmp_path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([None, "이름", "생년월일", "이메일"])
    sheet.append([None, "홍길동", datetime(1991, 3, 21), "hong@example.com"])
    file_path = tmp_path / "employees.xlsx"
    workbook.save(file_path)

    rows, skipped = read_excel_rows(str(file_path))

    assert skipped == 0
    assert rows[0].email == "hong@example.com"
    assert (rows[0].birth_month, rows[0].birth_day) == (3, 21)


def test_sync_hr_sheet_reads_configured_excel_path(tmp_path, monkeypatch) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["birthday", "name", "email"])
    sheet.append(["03-21", "홍길동", "hong@example.com"])
    file_path = tmp_path / "employees.xlsx"
    workbook.save(file_path)

    class FakeSlackClient:
        async def users_lookupByEmail(self, *, email: str) -> dict:
            assert email == "hong@example.com"
            return {"user": {"id": "U123"}}

    upserts = []

    async def upsert_birthday(pool, *, slack_user_id, birth_month, birth_day, email, source):
        upserts.append(
            {
                "slack_user_id": slack_user_id,
                "birth_month": birth_month,
                "birth_day": birth_day,
                "email": email,
                "source": source,
            }
        )

    async def mark_missing_birthdays_inactive(pool, active_slack_user_ids):
        assert active_slack_user_ids == ["U123"]
        return 0

    monkeypatch.setattr("sync.db.upsert_birthday", upsert_birthday)
    monkeypatch.setattr("sync.db.mark_missing_birthdays_inactive", mark_missing_birthdays_inactive)

    result = run(
        sync_hr_sheet(
            pool=object(),
            client=FakeSlackClient(),
            settings=SimpleNamespace(hr_excel_path=str(file_path)),
        )
    )

    assert result.upserted == 1
    assert result.deactivated == 0
    assert upserts == [
        {
            "slack_user_id": "U123",
            "birth_month": 3,
            "birth_day": 21,
            "email": "hong@example.com",
            "source": "hr",
        }
    ]


def test_sync_hr_sheet_returns_counts_before_lookup_abort(tmp_path, monkeypatch, caplog) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["birthday", "email"])
    sheet.append(["03-21", "ok@example.com"])
    sheet.append(["03-22", "missing@example.com"])
    sheet.append(["03-23", "fatal@example.com"])
    file_path = tmp_path / "employees.xlsx"
    workbook.save(file_path)

    class FakeSlackClient:
        async def users_lookupByEmail(self, *, email: str) -> dict:
            if email == "ok@example.com":
                return {"user": {"id": "U123"}}

            error = SlackApiError("lookup failed")
            error.response = {
                "error": "users_not_found" if email == "missing@example.com" else "fatal_error"
            }
            raise error

    upserts = []

    async def upsert_birthday(pool, *, slack_user_id, birth_month, birth_day, email, source):
        upserts.append((slack_user_id, source))

    async def mark_missing_birthdays_inactive(pool, active_slack_user_ids):
        raise AssertionError("soft-delete should be skipped on batch abort")

    monkeypatch.setattr("sync.db.upsert_birthday", upsert_birthday)
    monkeypatch.setattr("sync.db.mark_missing_birthdays_inactive", mark_missing_birthdays_inactive)

    result = run(
        sync_hr_sheet(
            pool=object(),
            client=FakeSlackClient(),
            settings=SimpleNamespace(hr_excel_path=str(file_path)),
        )
    )

    assert result.upserted == 1
    assert result.skipped == 1
    assert result.deactivated == 0
    assert upserts == [("U123", "hr")]
    assert "1 upserted, 1 skipped before batch abort" in caplog.text
